from __future__ import annotations

import json
import os
import re
import threading
import tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import BytesIO
from pathlib import Path
from typing import Any

import pypdfium2 as pdfium
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rapidocr_onnxruntime import RapidOCR


APP_VERSION = "20260608-railway-backend"
DEFAULT_POLLING_STATION = "1 - P Siddarampuram"
DEFAULT_SECTION_HEADING = "Section No and Name 1-MPP SCHOOL ROAD,SIDDARAMPURAM"

app = FastAPI(title="Voter Roll PDF to Excel API", version=APP_VERSION)
_ocr_local = threading.local()

allowed_origins = [
    origin.strip()
    for origin in os.getenv("ALLOWED_ORIGINS", "*").split(",")
    if origin.strip()
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=False,
    allow_methods=["GET", "POST", "OPTIONS"],
    allow_headers=["*"],
)


def bbox_stats(box: list[list[float]]) -> tuple[float, float, float, float, float, float]:
    xs = [point[0] for point in box]
    ys = [point[1] for point in box]
    x1, x2 = min(xs), max(xs)
    y1, y2 = min(ys), max(ys)
    return x1, y1, x2, y2, (x1 + x2) / 2, (y1 + y2) / 2


def clean_text(text: str) -> str:
    text = text.replace("|", "I")
    text = re.sub(r"\s+", " ", text).strip()
    text = text.replace("Name :", "Name:").replace("Name : ", "Name:")
    text = text.replace("Number :", "Number:")
    text = text.replace("Age :", "Age:")
    text = text.replace("Gender :", "Gender:")
    return text.strip()


def normalize_id(text: str) -> str | None:
    compact = re.sub(r"[^A-Z0-9]", "", text.upper())
    match = re.fullmatch(r"[A-Z]{3}\d{7}", compact)
    return match.group(0) if match else None


def parse_cell(lines: list[dict[str, Any]], voter_id: str) -> dict[str, str]:
    useful = []
    for line in sorted(lines, key=lambda item: (item["cy"], item["x1"])):
        text = clean_text(line["text"])
        compact = re.sub(r"\s+", "", text)
        if not text or text.lower() in {"photo", "available"}:
            continue
        if normalize_id(text) == voter_id:
            continue
        if re.fullmatch(r"\d{1,4}", compact):
            continue
        if "Photo" in text or "Available" in text:
            continue
        useful.append(text)

    name = ""
    relation_type = ""
    relation_name_parts: list[str] = []
    house_number = ""
    age = ""
    gender = ""
    mode = None

    for text in useful:
        if re.match(r"^Name\s*:", text, re.I):
            name = re.sub(r"^Name\s*:\s*", "", text, flags=re.I).strip()
            mode = "name"
            continue

        relation = re.match(
            r"^(Fathers|Husbands|Mothers|Others)\s*Name\s*:\s*(.*)$",
            text,
            re.I,
        )
        if relation:
            label = relation.group(1).lower()
            relation_type = {
                "fathers": "Father",
                "husbands": "Husband",
                "mothers": "Mother",
                "others": "Other",
            }.get(label, relation.group(1))
            if relation.group(2).strip():
                relation_name_parts = [relation.group(2).strip()]
            mode = "relation"
            continue

        house = re.match(r"^House\s*Number\s*:\s*(.*)$", text, re.I)
        if house:
            house_number = house.group(1).strip()
            mode = None
            continue

        age_gender = re.search(
            r"Age\s*:?\s*(\d{1,3})\s*Gend(?:er|ler)\s*:?\s*(Male|Female|M|F)",
            text,
            re.I,
        )
        if age_gender:
            age = age_gender.group(1)
            gender = "Male" if age_gender.group(2).lower().startswith("m") else "Female"
            mode = None
            continue

        if mode == "relation" and not re.search(r"House|Age|Gender|Gendler", text, re.I):
            relation_name_parts.append(text.strip())
        elif mode == "name" and not relation_type and not re.search(r"House|Age|Gender|Gendler", text, re.I):
            name = f"{name} {text}".strip()

    return {
        "Voter ID": voter_id,
        "Name": name,
        "Relation Type": relation_type,
        "Relation Name": " ".join(part for part in relation_name_parts if part).strip(),
        "House Number": house_number,
        "Age": age,
        "Gender": gender,
    }


def fixed_grid() -> tuple[list[float], float, list[float], float]:
    return [38, 566, 1095], 520, [75 + 220.5 * index for index in range(10)], 205


def cell_position(cx: float, cy: float, grid: tuple[list[float], float, list[float], float]) -> tuple[int, int]:
    col_lefts, col_width, row_tops, _ = grid
    col = min(range(len(col_lefts)), key=lambda index: abs(cx - (col_lefts[index] + col_width / 2)))
    row = min(range(len(row_tops)), key=lambda index: abs(cy - (row_tops[index] + 18)))
    return row, col


def cell_bounds(cx: float, cy: float, grid: tuple[list[float], float, list[float], float]) -> tuple[float, float, float, float]:
    col_lefts, col_width, row_tops, row_height = grid
    row, col = cell_position(cx, cy, grid)
    return (
        col_lefts[col] - 8,
        row_tops[row] - 10,
        col_lefts[col] + col_width + 8,
        row_tops[row] + row_height,
    )


def render_pdf_pages(pdf_path: Path, image_dir: Path, dpi: int) -> list[Path]:
    image_dir.mkdir(parents=True, exist_ok=True)
    pdf = pdfium.PdfDocument(pdf_path)
    scale = dpi / 72
    paths: list[Path] = []
    try:
        for index in range(len(pdf)):
            page = pdf[index]
            image = page.render(scale=scale).to_pil()
            image_path = image_dir / f"page-{index + 1:02d}.png"
            image.save(image_path)
            paths.append(image_path)
            page.close()
    finally:
        pdf.close()
    return paths


def ocr_image(image_path: Path) -> tuple[int, list[Any]]:
    page_number = int(re.search(r"page-(\d+)", image_path.name).group(1))
    if not hasattr(_ocr_local, "ocr"):
        _ocr_local.ocr = RapidOCR()
    ocr = _ocr_local.ocr
    result, _ = ocr(str(image_path))
    return page_number, result or []


def ocr_pages_parallel(image_paths: list[Path], max_workers: int) -> dict[int, list[Any]]:
    results: dict[int, list[Any]] = {}
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(ocr_image, path): path for path in image_paths}
        for future in as_completed(futures):
            page_number, page_result = future.result()
            results[page_number] = page_result
    return results


def extract_rows(
    ocr_results: dict[int, list[Any]],
    polling_station: str,
    start_page: int,
    end_page: int | None,
) -> list[dict[str, str]]:
    entries = []
    grid = fixed_grid()

    for page_num in sorted(ocr_results):
        if page_num < start_page or (end_page is not None and page_num > end_page):
            continue

        lines = []
        id_lines = []
        for box, text, conf in ocr_results[page_num]:
            x1, y1, x2, y2, cx, cy = bbox_stats(box)
            item = {
                "text": text,
                "conf": conf,
                "x1": x1,
                "y1": y1,
                "x2": x2,
                "y2": y2,
                "cx": cx,
                "cy": cy,
            }
            lines.append(item)
            voter_id = normalize_id(text)
            if voter_id:
                id_lines.append((voter_id, item))

        seen_ids = set()
        for voter_id, item in sorted(id_lines, key=lambda pair: (pair[1]["cy"], pair[1]["cx"])):
            if voter_id in seen_ids:
                continue
            seen_ids.add(voter_id)
            row_num, col_num = cell_position(item["cx"], item["cy"], grid)
            x1, y1, x2, y2 = cell_bounds(item["cx"], item["cy"], grid)
            cell_lines = [
                line
                for line in lines
                if x1 <= line["cx"] <= x2 and y1 <= line["cy"] <= y2
            ]
            parsed = parse_cell(cell_lines, voter_id)
            entries.append((page_num, row_num, col_num, parsed))

    entries.sort(key=lambda entry: (entry[0], entry[1], entry[2]))
    rows = []
    for index, (_, _, _, parsed) in enumerate(entries, start=1):
        row = {"No. and Name of Polling Station": polling_station, "Serial No": index}
        row.update(parsed)
        rows.append(row)
    return rows


def build_workbook(rows: list[dict[str, str]], section_heading: str) -> BytesIO:
    headers = [
        "No. and Name of Polling Station",
        "Serial No",
        "Voter ID",
        "Name",
        "Relation Type",
        "Relation Name",
        "House Number",
        "Age",
        "Gender",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Voter Roll"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"] = section_heading
    ws["A1"].font = Font(bold=True, color="7A3E00")
    ws["A1"].fill = PatternFill("solid", fgColor="FFF2CC")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{len(rows) + 2}"
    widths = [30, 10, 14, 34, 14, 36, 16, 8, 10]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def convert_pdf_to_workbook(
    pdf_path: Path,
    polling_station: str,
    section_heading: str,
    start_page: int,
    end_page: int | None,
    dpi: int,
    max_workers: int,
) -> tuple[BytesIO, dict[str, int]]:
    with tempfile.TemporaryDirectory(prefix="voter_roll_") as temp_dir:
        image_dir = Path(temp_dir) / "images"
        image_paths = render_pdf_pages(pdf_path, image_dir, dpi)
        ocr_results = ocr_pages_parallel(image_paths, max_workers=max_workers)
        rows = extract_rows(ocr_results, polling_station, start_page, end_page)
        workbook = build_workbook(rows, section_heading)
        missing = [
            row
            for row in rows
            if not row["Voter ID"]
            or not row["Name"]
            or not row["House Number"]
            or not row["Age"]
            or not row["Gender"]
        ]
        return workbook, {"entries": len(rows), "missing_key_fields": len(missing)}


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok", "version": APP_VERSION}


@app.post("/convert")
async def convert(
    file: UploadFile = File(...),
    polling_station: str = Form(DEFAULT_POLLING_STATION),
    section_heading: str = Form(DEFAULT_SECTION_HEADING),
    start_page: int = Form(3),
    end_page: int | None = Form(None),
    dpi: int = Form(200),
) -> StreamingResponse:
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Please upload a PDF file.")
    if start_page < 1:
        raise HTTPException(status_code=400, detail="start_page must be 1 or greater.")
    if end_page is not None and end_page < start_page:
        raise HTTPException(status_code=400, detail="end_page must be greater than or equal to start_page.")

    max_workers = max(1, min(int(os.getenv("OCR_WORKERS", "2")), 4))
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(await file.read())
        pdf_path = Path(tmp.name)

    try:
        workbook, stats = convert_pdf_to_workbook(
            pdf_path=pdf_path,
            polling_station=polling_station.strip() or DEFAULT_POLLING_STATION,
            section_heading=section_heading.strip() or DEFAULT_SECTION_HEADING,
            start_page=start_page,
            end_page=end_page,
            dpi=max(120, min(dpi, 260)),
            max_workers=max_workers,
        )
    finally:
        pdf_path.unlink(missing_ok=True)

    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", Path(file.filename).stem).strip("_") or "voter_roll"
    headers = {
        "Content-Disposition": f'attachment; filename="{safe_name}.xlsx"',
        "X-Total-Entries": str(stats["entries"]),
        "X-Missing-Key-Fields": str(stats["missing_key_fields"]),
        "X-Converter-Version": APP_VERSION,
    }
    return StreamingResponse(
        workbook,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
